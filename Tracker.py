import discord
from discord.ext import commands
import json
import os
from datetime import datetime, timedelta
import logging
from pathlib import Path
import time
from dotenv import load_dotenv
import pandas as pd
import re
from typing import List, Dict, Any, Optional, Set, Tuple
import tempfile
import asyncio
from collections import deque, defaultdict
import copy
import aiofiles
from functools import wraps
import hashlib
from urllib.parse import urlparse


# Load .env file
load_dotenv()

# Configuration
TRACKED_ROLE_IDS = os.getenv('TRACKED_ROLE_IDS', '').split(',') if os.getenv('TRACKED_ROLE_IDS') else []
TRACKED_ROLE_IDS = [role_id.strip() for role_id in TRACKED_ROLE_IDS if role_id.strip()]
BOT_OWNER_ID = os.getenv('BOT_OWNER_ID')
SECURITY_MANAGER_ROLE_ID = os.getenv('SECURITY_MANAGER_ROLE_ID')

# Setup logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('bot.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# Bot settings
intents = discord.Intents.default()
intents.message_content = True
intents.members = True
intents.guilds = True
intents.presences = True  # For member updates
bot = commands.Bot(command_prefix='!', intents=intents)

# File paths
BASE_DIR = Path(__file__).parent.absolute()
DATA_FILE = BASE_DIR / 'messages.json'
AUTHORIZED_CONFIG_FILE = BASE_DIR / 'authorized_config.json'

# Content channels management
CONTENT_CHANNELS_STORAGE_KEY = '__content_channels__'
LEGACY_CONTENT_CHANNELS_FILE = BASE_DIR / 'content_channels.json'
content_channels = set()

# In-memory cache for message data
message_cache: Dict[str, Any] = {}
message_cache_lock: Optional[asyncio.Lock] = None


def _is_meta_key(key: str) -> bool:
    return isinstance(key, str) and key.startswith('__')


def iter_user_entries(data: Dict[str, Any]):
    for user_id, user_data in data.items():
        if _is_meta_key(user_id):
            continue
        yield user_id, user_data


def _extract_content_channels(data: Dict[str, Any]) -> Tuple[Set[str], bool]:
    if not isinstance(data, dict):
        return set(), False

    storage = data.get(CONTENT_CHANNELS_STORAGE_KEY)
    if storage is None:
        return set(), False

    if isinstance(storage, dict):
        raw_channels = storage.get('channels', [])
    else:
        raw_channels = storage if isinstance(storage, list) else []

    channels = set()
    for channel_id in raw_channels:
        channel_str = str(channel_id).strip()
        if channel_str:
            channels.add(channel_str)
    return channels, True


# Authorization data
authorized_users: Set[str] = set()
authorized_roles: Set[str] = set()
authorization_lock = asyncio.Lock()


def get_message_cache_lock() -> asyncio.Lock:
    """Return the lock guarding the message cache, creating it on demand"""
    global message_cache_lock
    if message_cache_lock is None:
        message_cache_lock = asyncio.Lock()
    return message_cache_lock


def parse_authorization_entry(entry: str, default_type: str = 'user') -> Optional[Tuple[str, str]]:
    """Parse authorization entry into (type, id) tuple"""
    if not entry:
        return None

    value = entry.strip()
    if not value:
        return None

    entry_type = default_type
    lowered = value.lower()

    if lowered.startswith('role:'):
        entry_type = 'role'
        value = value.split(':', 1)[1].strip()
    elif lowered.startswith('user:'):
        entry_type = 'user'
        value = value.split(':', 1)[1].strip()

    if value.startswith('<@') and value.endswith('>'):
        inner = value[2:-1]
        if inner.startswith('&'):
            entry_type = 'role'
            inner = inner[1:]
        elif inner.startswith('!'):
            inner = inner[1:]
        value = inner

    value = value.strip()
    if not value:
        return None

    if value.lower().startswith('role:') or value.lower().startswith('user:'):
        return parse_authorization_entry(value)

    return entry_type, value

def load_content_channels():
    """Load content channels from the main messages data file"""
    global content_channels
    try:
        channels: Set[str] = set()
        found = False

        if message_cache:
            channels, found = _extract_content_channels(message_cache)
        elif DATA_FILE.exists():
            with open(DATA_FILE, 'r', encoding='utf-8') as f:
                data = json.load(f)
            channels, found = _extract_content_channels(data)

        if not found and LEGACY_CONTENT_CHANNELS_FILE.exists():
            with open(LEGACY_CONTENT_CHANNELS_FILE, 'r', encoding='utf-8') as f:
                legacy_data = json.load(f)
            legacy_channels = legacy_data.get('channels', []) if isinstance(legacy_data, dict) else legacy_data
            channels = {str(channel_id).strip() for channel_id in legacy_channels if str(channel_id).strip()}

        content_channels = channels
        logger.info(f"DEBUG: Loaded {len(content_channels)} content channels: {content_channels}")
    except Exception as e:
        logger.error(f"Error loading content channels: {e}")
        content_channels = set()

async def save_content_channels():
    """Persist content channels within the main messages data file"""
    try:
        await ensure_cache_loaded()
        async with get_message_cache_lock():
            storage = message_cache.setdefault(CONTENT_CHANNELS_STORAGE_KEY, {})
            if not isinstance(storage, dict):
                storage = {}
                message_cache[CONTENT_CHANNELS_STORAGE_KEY] = storage

            storage['channels'] = sorted(content_channels)
            storage['last_updated'] = datetime.now().isoformat()

        file_manager.mark_dirty()
        logger.info(f"DEBUG: Saved {len(content_channels)} content channels to messages data")
    except Exception as e:
        logger.error(f"Error saving content channels: {e}")


def load_authorization_from_env() -> None:
    """Load authorized users and roles from environment variable"""
    entries = os.getenv('AUTHORIZED_USERS', '')
    if not entries:
        return

    for raw_entry in entries.split(','):
        parsed = parse_authorization_entry(raw_entry)
        if not parsed:
            continue
        entry_type, identifier = parsed
        if entry_type == 'role':
            authorized_roles.add(identifier)
        else:
            authorized_users.add(identifier)


def load_authorization_from_file() -> None:
    """Load authorized users and roles from disk"""
    if not AUTHORIZED_CONFIG_FILE.exists():
        return

    try:
        with open(AUTHORIZED_CONFIG_FILE, 'r', encoding='utf-8') as f:
            data = json.load(f)

        for user_id in data.get('users', []):
            identifier = str(user_id).strip()
            if identifier:
                authorized_users.add(identifier)

        for role_id in data.get('roles', []):
            identifier = str(role_id).strip()
            if identifier:
                authorized_roles.add(identifier)

    except Exception as exc:
        logger.error(f"Failed to load authorization config: {exc}")


def save_authorization_config_sync() -> None:
    """Persist authorization configuration to disk"""
    try:
        data = {
            'users': sorted(authorized_users),
            'roles': sorted(authorized_roles)
        }
        temp_file = AUTHORIZED_CONFIG_FILE.with_suffix('.tmp')
        with open(temp_file, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
        temp_file.rename(AUTHORIZED_CONFIG_FILE)
    except Exception as exc:
        logger.error(f"Failed to save authorization config: {exc}")


def initialize_authorization() -> None:
    """Initialize authorization data from environment and file"""
    load_authorization_from_env()
    load_authorization_from_file()
    logger.info(
        "Authorization loaded: %s users, %s roles",
        len(authorized_users),
        len(authorized_roles)
    )


def is_bot_owner(member: discord.abc.User) -> bool:
    return bool(BOT_OWNER_ID) and str(getattr(member, 'id', None)) == BOT_OWNER_ID


def is_security_manager(member: discord.abc.User) -> bool:
    if not SECURITY_MANAGER_ROLE_ID:
        return False
    roles = getattr(member, 'roles', [])
    return any(str(role.id) == SECURITY_MANAGER_ROLE_ID for role in roles)


def is_authorized_member(member: discord.abc.User) -> bool:
    if str(getattr(member, 'id', None)) in authorized_users:
        return True
    roles = getattr(member, 'roles', [])
    return any(str(role.id) in authorized_roles for role in roles)


def has_full_access(member: discord.abc.User) -> bool:
    return is_bot_owner(member) or is_security_manager(member)

initialize_authorization()

async def add_authorized_identifier(entry_type: str, identifier: str) -> Tuple[bool, Optional[str], Optional[str]]:
    """Add an authorized user or role"""
    parsed = parse_authorization_entry(identifier, entry_type)
    if not parsed:
        return False, None, None
    resolved_type, normalized = parsed
    target_set = authorized_roles if resolved_type == 'role' else authorized_users

    async with authorization_lock:
        if normalized in target_set:
            return False, normalized, resolved_type
        target_set.add(normalized)
        save_authorization_config_sync()

    return True, normalized, resolved_type


async def remove_authorized_identifier(entry_type: str, identifier: str) -> Tuple[bool, Optional[str], Optional[str]]:
    """Remove an authorized user or role"""
    parsed = parse_authorization_entry(identifier, entry_type)
    if not parsed:
        return False, None, None
    resolved_type, normalized = parsed
    target_set = authorized_roles if resolved_type == 'role' else authorized_users

    async with authorization_lock:
        if normalized not in target_set:
            return False, normalized, resolved_type
        target_set.remove(normalized)
        save_authorization_config_sync()

    return True, normalized, resolved_type


async def get_authorization_snapshot() -> Tuple[List[str], List[str]]:
    """Return copies of authorized user and role lists"""
    async with authorization_lock:
        with_ids = sorted(authorized_users)
        with_roles = sorted(authorized_roles)
    return with_ids, with_roles


PUBLIC_COMMANDS: Set[str] = {
    'helpme',
    'trackcommands',
    'stats',
    'system_info',
    'user_info',
    'top_users',
    'listcontentchannels',
    'api_stats'
}


@bot.check
async def global_permission_check(ctx: commands.Context) -> bool:
    command = ctx.command.name if ctx.command else None
    if command is None:
        return True

    if has_full_access(ctx.author):
        return True

    if command in PUBLIC_COMMANDS:
        return True

    if command == 'filter':
        return is_authorized_member(ctx.author)

    return False


def is_content_channel(channel_id: str) -> bool:
    """Check if a channel is a content channel"""
    # Ensure content channels are loaded
    global content_channels
    if not content_channels:
        load_content_channels()

    # Convert both to string for consistent comparison
    channel_id_str = str(channel_id)
    result = channel_id_str in content_channels
    return result

def should_track_user(member) -> bool:
    """Check if a user should be tracked based on their roles"""
    if not TRACKED_ROLE_IDS:
        # If no roles specified, track everyone (current behavior)
        return True

    # Check if user has any of the tracked roles
    user_role_ids = [str(role.id) for role in member.roles]
    return any(role_id in user_role_ids for role_id in TRACKED_ROLE_IDS)

def migrate_user_data(user_data: Dict[str, Any]) -> Dict[str, Any]:
    """Migrate old messages format to new counter format"""
    if 'messages' in user_data and 'counter' not in user_data:
        # Old format detected, convert to new format
        messages = user_data['messages']
        counter = SmartCounter()

        for msg in messages:
            if isinstance(msg, dict):
                timestamp = msg.get('timestamp', datetime.now().isoformat())
                channel_name = msg.get('channel_name', 'unknown')
                counter.add_message(timestamp, channel_name)

        # Replace messages with counter
        migrated_data = user_data.copy()
        del migrated_data['messages']
        migrated_data['counter'] = counter.to_dict()

        logger.info(f"Migrated user data from old format (messages: {len(messages)})")
        return migrated_data

    return user_data

def detect_x_links(text: str) -> List[str]:
    """Detect X (Twitter) links in text"""
    # X (Twitter) link patterns - updated to catch more variations
    x_patterns = [
        r'https?://(?:twitter\.com|x\.com)/[a-zA-Z0-9_]+/status/[0-9]+',
        r'https?://(?:twitter\.com|x\.com)/[a-zA-Z0-9_]+',
        r'https?://(?:www\.)?(?:twitter\.com|x\.com)/[^\s]+'
    ]

    all_links = []
    for pattern in x_patterns:
        links = re.findall(pattern, text)
        all_links.extend(links)

    # Remove duplicates while preserving order - prioritize complete URLs
    unique_links = []
    seen = set()

    # First pass: collect all links
    for link in all_links:
        if link not in seen:
            seen.add(link)
            unique_links.append(link)

    # Second pass: filter out partial URLs that are substrings of more complete URLs
    filtered_links = []
    for i, link in enumerate(unique_links):
        is_partial = False
        for j, other_link in enumerate(unique_links):
            if i != j and link in other_link and len(link) < len(other_link):
                # This link is a substring of a longer link, so it's a partial match
                logger.debug(f"DEBUG: Removing partial link '{link}' because it's contained in '{other_link}'")
                is_partial = True
                break

        if not is_partial:
            filtered_links.append(link)

    # Clean up links (remove trailing punctuation)
    cleaned_links = []
    for link in filtered_links:
        # Remove trailing characters like .,!?;:)
        cleaned_link = link.rstrip('.,!?;:)')
        cleaned_links.append(cleaned_link)
    return cleaned_links

async def process_content_links(
    user_id: str,
    message_content: str,
    timestamp: str
) -> Tuple[List[str], List[str]]:
    """Process content links in a message.

    Returns tuple of (all detected links, newly stored unique links).
    """
    detected_links = detect_x_links(message_content)

    # Debug: Log detected links
    if detected_links:
        await ensure_cache_loaded()
        async with get_message_cache_lock():
            if user_id not in message_cache:
                message_cache[user_id] = {
                    'username': 'Unknown',
                    'roles': [],
                    'counter': SmartCounter().to_dict(),
                    'content_counter': ContentCounter().to_dict()
                }
            elif 'content_counter' not in message_cache[user_id]:
                message_cache[user_id]['content_counter'] = ContentCounter().to_dict()

            content_counter = ContentCounter.from_dict(message_cache[user_id]['content_counter'])

            new_links: List[str] = []
            for link in detected_links:
                is_new_link = content_counter.add_content(timestamp, link)
                if is_new_link:
                    new_links.append(link)
                    logger.info(f"DEBUG: Added new content link: {link}")
                else:
                    logger.info(f"DEBUG: Skipped duplicate content link: {link}")

            message_cache[user_id]['content_counter'] = content_counter.to_dict()

        if new_links:
            file_manager.mark_dirty()
            logger.info(
                "Added %s new content links for user %s (detected: %s)",
                len(new_links),
                user_id,
                len(detected_links)
            )

        return detected_links, new_links

    return [], []
# Optimized data structure - only essential data
tracker_stats = {
    'messages_processed': 0,
    'start_time': time.time(),
    'api_calls': 0,
    'rate_limit_hits': 0,
    'disk_writes': 0,
    'cache_hits': 0
}

# UserCache class removed - using direct JSON access for better performance

async def get_message_cache_snapshot() -> Dict[str, Any]:
    """Create a deep copy snapshot of the in-memory message cache"""
    async with get_message_cache_lock():
        return copy.deepcopy(message_cache)

class RateLimiter:
    """Rate limit handler for Discord API"""
    
    def __init__(self, max_calls: int = 45, window: int = 60):
        self.max_calls = max_calls
        self.window = window
        self.calls = deque()
        self.lock = asyncio.Lock()
    
    async def acquire(self) -> None:
        """Acquire rate limit slot"""
        async with self.lock:
            now = time.time()
            
            # Remove old calls
            while self.calls and self.calls[0] <= now - self.window:
                self.calls.popleft()
            
            # Check if we're at limit
            if len(self.calls) >= self.max_calls:
                # Calculate wait time
                wait_time = self.calls[0] + self.window - now
                if wait_time > 0:
                    tracker_stats['rate_limit_hits'] += 1
                    logger.warning(f"Rate limit hit, waiting {wait_time:.2f}s")
                    await asyncio.sleep(wait_time)
            
            # Add current call
            self.calls.append(now)
            tracker_stats['api_calls'] += 1

class AsyncFileManager:
    """Async file manager that batches writes to reduce I/O pressure"""

    def __init__(self, flush_interval: float = 2.0):
        self.flush_interval = flush_interval
        self.is_running = False
        self.worker_task = None
        self.dirty = False
        self.last_dirty_time = 0.0
        self._flush_lock = asyncio.Lock()

    async def start(self) -> None:
        """Start the background flush worker"""
        if not self.is_running:
            self.is_running = True
            self.worker_task = asyncio.create_task(self._worker())
            logger.info("Async file manager started")

    async def stop(self) -> None:
        """Stop the worker and flush pending data"""
        if self.is_running:
            self.is_running = False
            if self.worker_task:
                await self.flush_now()
                self.worker_task.cancel()
                try:
                    await self.worker_task
                except asyncio.CancelledError:
                    pass
            logger.info("Async file manager stopped")

    def mark_dirty(self) -> None:
        """Mark the cache as dirty to trigger a future flush"""
        self.dirty = True
        self.last_dirty_time = time.monotonic()

    async def flush_now(self) -> None:
        """Flush immediately if there are pending changes"""
        if not self.dirty:
            return
        async with self._flush_lock:
            if not self.dirty:
                return
            snapshot = await get_message_cache_snapshot()
            await self._write_to_file(snapshot)
            self.dirty = False

    async def _worker(self) -> None:
        """Periodic worker that batches writes within the flush interval"""
        try:
            while self.is_running:
                if self.dirty:
                    elapsed = time.monotonic() - self.last_dirty_time
                    remaining = self.flush_interval - elapsed
                    if remaining > 0:
                        await asyncio.sleep(min(remaining, self.flush_interval))
                        continue
                    await self.flush_now()
                else:
                    await asyncio.sleep(0.5)
        except asyncio.CancelledError:
            raise
        except Exception as e:
            logger.error(f"Error in file worker: {e}")

    async def _write_to_file(self, data: Dict[str, Any]) -> None:
        """Write data to file asynchronously"""
        try:
            # Debug: Check data type
            if not isinstance(data, dict):
                logger.error(f"Invalid data type for writing: {type(data)}. Expected dict.")
                return

            # Debug: Log before json.dumps
            logger.debug(f"Attempting to serialize data with {len(data)} users")

            # Write to temporary file first
            temp_file = DATA_FILE.with_suffix('.tmp')
            async with aiofiles.open(temp_file, 'w', encoding='utf-8') as f:
                json_str = json.dumps(data, indent=2, ensure_ascii=False, default=self._json_default_handler)
                await f.write(json_str)

            # Atomic rename
            temp_file.rename(DATA_FILE)
            tracker_stats['disk_writes'] += 1

        except Exception as e:
            logger.error(f"Error writing to file: {e}")
            logger.error(f"Data type: {type(data)}")
            # Try to find the problematic object
            for user_id, user_data in data.items():
                if _is_meta_key(user_id):
                    continue
                if not isinstance(user_data, dict):
                    logger.error(f"Problematic user {user_id}: {type(user_data)} = {user_data}")
                else:
                    for key, value in user_data.items():
                        if hasattr(value, '__dict__') and 'Command' in str(type(value)):
                            logger.error(f"Found Command object in user {user_id}, key {key}: {type(value)}")
                        if isinstance(value, (list, dict)):
                            self._debug_recursive(value, f"user_{user_id}_{key}")

    def _debug_recursive(self, obj, path=""):
        """Recursively debug objects to find non-serializable items"""
        if isinstance(obj, dict):
            for key, value in obj.items():
                if hasattr(value, '__dict__'):
                    obj_type = str(type(value))
                    if 'Command' in obj_type:
                        logger.error(f"Found Command object at {path}.{key}: {obj_type}")
                        logger.error(f"Command object details: {value}")
                elif isinstance(value, (dict, list)):
                    self._debug_recursive(value, f"{path}.{key}")
        elif isinstance(obj, list):
            for i, item in enumerate(obj):
                if hasattr(item, '__dict__'):
                    obj_type = str(type(item))
                    if 'Command' in obj_type:
                        logger.error(f"Found Command object at {path}[{i}]: {obj_type}")
                        logger.error(f"Command object details: {item}")
                elif isinstance(item, (dict, list)):
                    self._debug_recursive(item, f"{path}[{i}]")

    def _json_default_handler(self, obj):
        """Custom JSON serializer for non-serializable objects"""
        if hasattr(obj, '__dict__'):
            obj_type = str(type(obj))
            if 'Command' in obj_type:
                logger.warning(f"Found Command object during JSON serialization: {obj_type}")
                return f"<{obj_type}>"
        return str(obj)

# Initialize systems
rate_limiter = RateLimiter()
file_manager = AsyncFileManager()

async def get_user_info_from_json(user_id: str) -> Optional[Dict[str, Any]]:
    """Get user info directly from JSON file"""
    try:
        await ensure_cache_loaded()

        async with get_message_cache_lock():
            if user_id in message_cache:
                user = message_cache[user_id]
                return {
                    'username': user.get('username', 'Unknown'),
                    'roles': user.get('roles', [])
                }
        return None
    except Exception as e:
        logger.error(f"Error reading user info from JSON: {e}")
        return None

async def migrate_user_role_ids():
    """Migrate existing users to add role_ids field"""
    try:
        logger.info("Starting migration of user role IDs...")

        await ensure_cache_loaded()

        async with get_message_cache_lock():
            data = copy.deepcopy(message_cache)

        migrated_count = 0
        # Get guild and update all users with their role IDs
        guild = bot.guilds[0] if bot.guilds else None

        if guild:
            for user_id, user_data in data.items():
                if _is_meta_key(user_id):
                    continue
                # Skip if already has role_ids
                if 'role_ids' in user_data:
                    continue

                # Get member from guild
                member = guild.get_member(int(user_id))
                if member:
                    # Extract role IDs
                    role_ids = [str(role.id) for role in member.roles if role.name != '@everyone']
                    user_data['role_ids'] = role_ids
                    migrated_count += 1

                    logger.debug(f"Migrated role IDs for user {user_id}: {role_ids}")
                else:
                    logger.warning(f"Member not found for user {user_id}")

            # Save updated data
            async with get_message_cache_lock():
                message_cache.clear()
                message_cache.update(data)
            file_manager.mark_dirty()
            logger.info(f"Migration completed. Updated {migrated_count} users with role IDs.")
        else:
            logger.error("No guild found for migration")

    except Exception as e:
        logger.error(f"Error during role ID migration: {e}")

async def update_user_in_json(user_id: str, username: str, roles: List[str], role_ids: List[str] = None) -> None:
    """Update user info in JSON file"""
    try:
        await ensure_cache_loaded()
        async with get_message_cache_lock():
            if user_id not in message_cache:
                message_cache[user_id] = {
                    'username': username,
                    'roles': roles,
                    'counter': SmartCounter().to_dict()
                }

            message_cache[user_id]['username'] = username
            message_cache[user_id]['roles'] = roles
            if role_ids:
                message_cache[user_id]['role_ids'] = role_ids

        file_manager.mark_dirty()

    except Exception as e:
        logger.error(f"Error updating user info in JSON: {e}")

def async_task(coro):
    """Decorator to run async tasks in background"""
    @wraps(coro)
    def wrapper(*args, **kwargs):
        return asyncio.create_task(coro(*args, **kwargs))
    return wrapper

class SmartCounter:
    """Intelligent message counting system with 24-hour rolling window and channel-specific tracking"""

    def __init__(self):
        self.total_count = 0
        self.daily_counts = defaultdict(int)
        self.channel_counts = defaultdict(int)
        self.message_timestamps = []  # Store all message timestamps for accurate 24h counting
        self.channel_timestamps = defaultdict(list)  # Store timestamps per channel for accurate channel-specific counting
        self.first_seen = None
        self.last_seen = None

    def add_message(self, timestamp: str, channel_name: str) -> None:
        """Add a message to the counter"""
        # Extract date from timestamp for daily counts
        date = timestamp.split('T')[0]

        # Update counts
        self.total_count += 1
        self.daily_counts[date] += 1
        self.channel_counts[channel_name] += 1

        # Store timestamp for accurate 24-hour counting
        self.message_timestamps.append(timestamp)

        # Store timestamp for channel-specific counting
        self.channel_timestamps[channel_name].append(timestamp)

        # Keep only last 30 days of timestamps for performance
        self._cleanup_old_timestamps()
        self._cleanup_channel_timestamps()

        # Update timestamps
        if not self.first_seen:
            self.first_seen = timestamp
        self.last_seen = timestamp

    def _cleanup_old_timestamps(self, keep_days: int = 30) -> None:
        """Clean up old timestamps to prevent memory issues"""
        if not self.message_timestamps:
            return

        cutoff_time = datetime.now() - timedelta(days=keep_days)
        cutoff_iso = cutoff_time.isoformat()

        # Keep only recent timestamps
        self.message_timestamps = [
            ts for ts in self.message_timestamps
            if ts >= cutoff_iso
        ]

    def _cleanup_channel_timestamps(self, keep_days: int = 30) -> None:
        """Clean up old channel timestamps to prevent memory issues"""
        cutoff_time = datetime.now() - timedelta(days=keep_days)
        cutoff_iso = cutoff_time.isoformat()

        # Clean up timestamps for each channel
        for channel_name in self.channel_timestamps:
            self.channel_timestamps[channel_name] = [
                ts for ts in self.channel_timestamps[channel_name]
                if ts >= cutoff_iso
            ]

    def get_time_range_count(self, days: int) -> int:
        """Get message count for last N days with accurate 24-hour calculation"""
        if days <= 0:
            return self.total_count

        if days == 1:  # Special case for 24 hours (1 day)
            return self._get_24h_count()

        # For multi-day periods, use daily counts
        cutoff_date = (datetime.now() - timedelta(days=days)).strftime('%Y-%m-%d')
        return sum(
            count for date, count in self.daily_counts.items()
            if date >= cutoff_date
        )

    def _get_24h_count(self) -> int:
        """Get accurate message count for last 24 hours"""
        if not self.message_timestamps:
            return 0

        # Calculate cutoff time (24 hours ago from now)
        cutoff_time = datetime.now() - timedelta(hours=24)
        cutoff_iso = cutoff_time.isoformat()

        # Count messages in the last 24 hours
        count = sum(1 for ts in self.message_timestamps if ts >= cutoff_iso)
        return count

    def get_hourly_count(self, hours: int) -> int:
        """Get message count for last N hours"""
        if hours <= 0:
            return self.total_count

        if not self.message_timestamps:
            return 0

        # Calculate cutoff time
        cutoff_time = datetime.now() - timedelta(hours=hours)
        cutoff_iso = cutoff_time.isoformat()

        # Count messages in the last N hours
        count = sum(1 for ts in self.message_timestamps if ts >= cutoff_iso)
        return count

    def get_channel_count_hours(self, channel_name: str, hours: int) -> int:
        """Get message count for specific channel in the last N hours"""
        if hours <= 0:
            return self.channel_counts.get(channel_name, 0)

        if channel_name not in self.channel_timestamps:
            return 0

        cutoff_time = datetime.now() - timedelta(hours=hours)
        cutoff_iso = cutoff_time.isoformat()

        return sum(1 for ts in self.channel_timestamps[channel_name] if ts >= cutoff_iso)

    def get_channel_count(self, channel_name: str, days: int = None) -> int:
        """Get message count for specific channel with accurate time-based calculation"""
        logger.info(f"DEBUG: get_channel_count called for channel '{channel_name}', days={days}")

        if days is None:
            count = self.channel_counts.get(channel_name, 0)
            logger.info(f"DEBUG: No time range, returning total channel count: {count}")
            return count

        # For time ranges, use accurate timestamp-based calculation
        if days <= 0:
            count = self.channel_counts.get(channel_name, 0)
            logger.info(f"DEBUG: Invalid days ({days}), returning total channel count: {count}")
            return count

        if days == 1:  # 24 hours
            count = self._get_channel_24h_count(channel_name)
            logger.info(f"DEBUG: 24h count for '{channel_name}': {count}")
            return count

        # For multi-day periods, use channel-specific timestamps
        count = self._get_channel_time_range_count(channel_name, days)
        logger.info(f"DEBUG: {days} day count for '{channel_name}': {count}")
        return count

    def _get_channel_24h_count(self, channel_name: str) -> int:
        """Get accurate message count for specific channel in last 24 hours"""
        logger.info(f"DEBUG: _get_channel_24h_count called for channel '{channel_name}'")
        logger.info(f"DEBUG: Available channels in channel_timestamps: {list(self.channel_timestamps.keys())}")

        if channel_name not in self.channel_timestamps:
            logger.info(f"DEBUG: Channel '{channel_name}' not found in channel_timestamps, returning 0")
            return 0

        timestamps = self.channel_timestamps[channel_name]
        logger.info(f"DEBUG: Channel '{channel_name}' has {len(timestamps)} timestamps")

        # Calculate cutoff time (24 hours ago from now)
        cutoff_time = datetime.now() - timedelta(hours=24)
        cutoff_iso = cutoff_time.isoformat()
        logger.info(f"DEBUG: 24h cutoff time: {cutoff_iso}")

        # Count messages in the last 24 hours for this channel
        count = sum(1 for ts in timestamps if ts >= cutoff_iso)
        logger.info(f"DEBUG: Found {count} messages in last 24h for channel '{channel_name}'")
        return count

    def _get_channel_time_range_count(self, channel_name: str, days: int) -> int:
        """Get message count for specific channel in last N days"""
        if channel_name not in self.channel_timestamps:
            return 0

        # Calculate cutoff time
        cutoff_time = datetime.now() - timedelta(days=days)
        cutoff_iso = cutoff_time.isoformat()

        # Count messages in the last N days for this channel
        count = sum(1 for ts in self.channel_timestamps[channel_name] if ts >= cutoff_iso)
        return count

    def cleanup_old_data(self, keep_days: int = 365) -> None:
        """Clean up old daily count data"""
        cutoff_date = (datetime.now() - timedelta(days=keep_days)).strftime('%Y-%m-%d')

        old_dates = [date for date in self.daily_counts.keys() if date < cutoff_date]
        for date in old_dates:
            del self.daily_counts[date]

    def to_dict(self) -> Dict[str, Any]:
        """Convert to dictionary for JSON serialization"""
        return {
            'total_count': self.total_count,
            'daily_counts': dict(self.daily_counts),
            'channel_counts': dict(self.channel_counts),
            'message_timestamps': self.message_timestamps,
            'channel_timestamps': dict(self.channel_timestamps),
            'first_seen': self.first_seen,
            'last_seen': self.last_seen
        }

    @classmethod
    def from_dict(cls, data: Dict[str, Any]) -> 'SmartCounter':
        """Create from dictionary with migration support"""
        counter = cls()
        counter.total_count = data.get('total_count', 0)
        counter.daily_counts = defaultdict(int, data.get('daily_counts', {}))
        counter.channel_counts = defaultdict(int, data.get('channel_counts', {}))
        counter.message_timestamps = data.get('message_timestamps', [])
        counter.channel_timestamps = defaultdict(list, data.get('channel_timestamps', {}))
        counter.first_seen = data.get('first_seen')
        counter.last_seen = data.get('last_seen')

        # Migrate old data: if no message_timestamps but we have daily_counts,
        # generate some representative timestamps
        if not counter.message_timestamps and counter.daily_counts:
            logger.info(f"Migrating SmartCounter: generating timestamps for {counter.total_count} messages")
            # This is a best-effort migration - distribute timestamps evenly
            counter.message_timestamps = []
            now = datetime.now()
            for date, count in counter.daily_counts.items():
                # Create timestamps for this date, distributed evenly
                date_obj = datetime.strptime(date, '%Y-%m-%d')
                for i in range(count):
                    # Distribute messages throughout the day
                    hour = (i * 24) // max(count, 1)  # Spread across 24 hours
                    minute = ((i * 24 * 60) // max(count, 1)) % 60
                    timestamp = date_obj.replace(hour=hour, minute=minute).isoformat()
                    counter.message_timestamps.append(timestamp)

            # Sort timestamps and keep only recent ones
            counter.message_timestamps.sort()
            counter._cleanup_old_timestamps()
            logger.info(f"Migration completed: {len(counter.message_timestamps)} timestamps generated")

        # Migrate channel timestamps if they don't exist
        if not counter.channel_timestamps and counter.message_timestamps:
            logger.info(f"Migrating SmartCounter: generating channel timestamps")
            # Distribute existing timestamps across channels based on channel_counts
            for channel_name, channel_count in counter.channel_counts.items():
                if channel_count > 0:
                    # Calculate ratio of messages for this channel
                    ratio = channel_count / counter.total_count
                    channel_timestamp_count = int(len(counter.message_timestamps) * ratio)

                    # Take that many timestamps from the end (most recent)
                    if channel_timestamp_count > 0:
                        counter.channel_timestamps[channel_name] = counter.message_timestamps[-channel_timestamp_count:]

            logger.info(f"Channel migration completed: {len(counter.channel_timestamps)} channels populated")

        return counter

class ContentCounter:
    """Content link counting system for X (Twitter) links"""
    
    def __init__(self):
        self.total_count = 0
        self.daily_counts = defaultdict(int)  # Daily content counts
        self.content_links = []  # List of content links with timestamps (for duplicate checking and export)
        self.normalized_links = set()  # Set for quick duplicate checking
        self.first_content_seen = None
        self.last_content_seen = None
    
    def normalize_url(self, url: str) -> str:
        """Normalize URL for duplicate checking"""
        try:
            # Parse URL
            parsed = urlparse(url.lower())

            # Remove www. from domain
            domain = parsed.netloc
            if domain.startswith('www.'):
                domain = domain[4:]

            # Reconstruct URL with consistent format
            normalized = f"{parsed.scheme}://{domain}{parsed.path}"
            if parsed.query:
                normalized += f"?{parsed.query}"
            if parsed.fragment:
                normalized += f"#{parsed.fragment}"

            return normalized
        except:
            # If parsing fails, return lowercase original
            return url.lower()

    def add_content(self, timestamp: str, link: str) -> bool:
        """Add a content link and update counts.

        Returns True when the link is stored for the first time, False otherwise.
        """
        # Check for duplicates
        normalized_link = self.normalize_url(link)

        # Debug: Log normalization process
        logger.debug(f"DEBUG: Normalizing link: {link} -> {normalized_link}")
        logger.debug(f"DEBUG: Current normalized_links: {self.normalized_links}")

        # Check if this link already exists (using set for O(1) lookup)
        if normalized_link in self.normalized_links:
            logger.debug(f"Duplicate content link skipped: {link}")
            return False  # Skip adding duplicate

        # Add link to list (for export and duplicate checking)
        self.content_links.append({
            'link': link,
            'timestamp': timestamp
        })

        # Add to normalized links set
        self.normalized_links.add(normalized_link)

        # Debug: Log successful addition
        logger.debug(f"DEBUG: Added new content link: {link}")
        logger.debug(f"DEBUG: New total_count: {self.total_count + 1}")

        # Update counts (SmartCounter style)
        date = timestamp.split('T')[0]  # Extract date from timestamp
        self.total_count += 1
        self.daily_counts[date] += 1

        # Update timestamps
        if not self.first_content_seen:
            self.first_content_seen = timestamp
        self.last_content_seen = timestamp
        return True
    
    def get_time_range_count(self, days: int) -> int:
        """Get content count for last N days (same as SmartCounter)"""
        if days <= 0:
            return self.total_count
        cutoff_date = (datetime.now() - timedelta(days=days)).strftime('%Y-%m-%d')
        return sum(
            count for date, count in self.daily_counts.items()
            if date >= cutoff_date
        )

    def get_time_range_count_hours(self, hours: int) -> int:
        """Get content count for the last N hours"""
        if hours <= 0:
            return self.total_count

        cutoff_time = datetime.now() - timedelta(hours=hours)

        def parse_timestamp(ts: str) -> datetime:
            try:
                return datetime.fromisoformat(ts)
            except ValueError:
                return datetime.fromisoformat(ts.replace('Z', '+00:00'))

        return sum(
            1
            for link in self.content_links
            if parse_timestamp(link['timestamp']) >= cutoff_time
        )
    
    def get_total_content_count(self) -> int:
        """Get total content links count"""
        return self.total_count
    
    def cleanup_old_links(self, keep_days: int = 365) -> None:
        """Clean up old link data"""
        cutoff_date = datetime.now() - timedelta(days=keep_days)

        old_links = [link for link in self.content_links
                    if datetime.fromisoformat(link['timestamp'].replace('Z', '+00:00')) < cutoff_date]

        for link in old_links:
            self.content_links.remove(link)

        # Rebuild normalized_links set after cleanup
        self.normalized_links = set()
        for link_data in self.content_links:
            self.normalized_links.add(self.normalize_url(link_data['link']))

        # Clean up old daily counts
        cutoff_date_str = cutoff_date.strftime('%Y-%m-%d')
        old_dates = [date for date in self.daily_counts.keys() if date < cutoff_date_str]
        for date in old_dates:
            del self.daily_counts[date]

        # Recalculate total count
        self.total_count = sum(self.daily_counts.values())
    
    def to_dict(self) -> Dict[str, Any]:
        """Convert to dictionary for JSON serialization"""
        return {
            'total_count': self.total_count,
            'daily_counts': dict(self.daily_counts),
            'content_links': self.content_links,
            'first_content_seen': self.first_content_seen,
            'last_content_seen': self.last_content_seen
        }
    
    @classmethod
    def from_dict(cls, data: Dict[str, Any]) -> 'ContentCounter':
        """Create from dictionary with migration support"""
        counter = cls()

        # Check if this is old format data (has content_counts but no daily_counts)
        if 'content_counts' in data and 'daily_counts' not in data:
            # Migrate from old format to new format
            logger.info("Migrating ContentCounter from old format to new format")

            counter.content_links = data.get('content_links', [])
            counter.first_content_seen = data.get('first_content_seen')
            counter.last_content_seen = data.get('last_content_seen')

            # Calculate daily_counts from content_links
            counter.daily_counts = defaultdict(int)
            for link_data in counter.content_links:
                date = link_data['timestamp'].split('T')[0]
                counter.daily_counts[date] += 1

            # Set total_count
            counter.total_count = len(counter.content_links)

            logger.info(f"Migrated {counter.total_count} content links with {len(counter.daily_counts)} daily entries")

        else:
            # New format data
            counter.total_count = data.get('total_count', 0)
            counter.daily_counts = defaultdict(int, data.get('daily_counts', {}))
            counter.content_links = data.get('content_links', [])
            counter.first_content_seen = data.get('first_content_seen')
            counter.last_content_seen = data.get('last_content_seen')

        # Rebuild normalized_links set for duplicate checking
        counter.normalized_links = set()
        for link_data in counter.content_links:
            counter.normalized_links.add(counter.normalize_url(link_data['link']))

        return counter

class FilterEngine:
    """Optimized filtering engine for user data"""
    
    def __init__(self):
        self.operators = ['>', '<', '>=', '<=', '==', '!=']
        self.time_units = {'h': 'hours', 'd': 'days', 'w': 'weeks', 'm': 'months', 'y': 'years'}
    
    def parse_filter_command(self, command_text: str) -> Dict[str, Any]:
        """Parse filter command like 'message>10 @rol1 and @rol2 #kanal 7d'"""
        filters = {
            'message_count': None,
            'roles': {'include': [], 'exclude': [], 'operator': 'and'},
            'channels': [],
            'time_range': None,
            'raw_command': command_text
        }
        
        # Extract message count filter
        message_match = re.search(r'message([><=!]+)(\d+)', command_text)
        if message_match:
            operator, count = message_match.groups()
            filters['message_count'] = {'operator': operator, 'value': int(count)}
        
        # Extract content count filter
        content_match = re.search(r'content([><=!]+)(\d+)', command_text)
        if content_match:
            operator, count = content_match.groups()
            filters['content_count'] = {'operator': operator, 'value': int(count)}
        
        # Extract role filters and nohave filters
        # Split command by nohave to separate include/exclude roles
        nohave_parts = command_text.split(' nohave ', 1)
        include_command = nohave_parts[0]
        exclude_command = nohave_parts[1] if len(nohave_parts) > 1 else ''

        # Process include roles
        role_pattern = r'@([^@#\d]+?)(?=\s*(?:@|#|\band\b|\bor\b|\bnohave\b|\d+[dwmy]|$))|<@&([0-9]+)>'
        include_matches = re.findall(role_pattern, include_command)

        # Process exclude roles (after nohave)
        exclude_matches = []
        if exclude_command:
            exclude_matches = re.findall(role_pattern, exclude_command)
        
        # Clean up include role names and IDs
        cleaned_include_roles = []
        for match in include_matches:
            # match is a tuple: (role_name, role_id) - one will be empty
            role_name, role_id = match
            if role_name:
                role_name = role_name.strip()
                # Remove trailing "and" or "or" if they were captured
                if role_name.lower().endswith(' and'):
                    role_name = role_name[:-4].strip()
                elif role_name.lower().endswith(' or'):
                    role_name = role_name[:-3].strip()
                if role_name:
                    cleaned_include_roles.append(role_name)
            elif role_id:
                cleaned_include_roles.append(f"<@&{role_id}>")

        # Clean up exclude role names and IDs
        cleaned_exclude_roles = []
        for match in exclude_matches:
            # match is a tuple: (role_name, role_id) - one will be empty
            role_name, role_id = match
            if role_name:
                role_name = role_name.strip()
                if role_name:
                    cleaned_exclude_roles.append(role_name)
            elif role_id:
                cleaned_exclude_roles.append(f"<@&{role_id}>")
        
        if cleaned_include_roles:
            # Check for AND/OR logic (only in include part)
            if ' and ' in include_command.lower():
                filters['roles']['operator'] = 'and'
            elif ' or ' in include_command.lower():
                filters['roles']['operator'] = 'or'
            else:
                filters['roles']['operator'] = 'or'  # Default to OR if not specified
            filters['roles']['include'] = cleaned_include_roles

        if cleaned_exclude_roles:
            filters['roles']['exclude'] = cleaned_exclude_roles
        
        # Extract channel filters (support both #channel and <#ID> formats)
        channel_matches = re.findall(r'<#([0-9]+)>|#([^\s]+)', command_text)
        logger.info(f"Channel matches found: {channel_matches}")
        logger.info(f"Original command text: {command_text}")

        # Process matches to get channel names/IDs
        channels = []
        for match in channel_matches:
            # match is a tuple: (id, name) - one will be empty
            channel_id, channel_name = match
            if channel_id:
                channels.append(channel_id)
            elif channel_name:
                channels.append(channel_name)

        if channels:
            # Convert channel IDs to names if needed
            processed_channels = []
            for channel in channels:
                # If it's a numeric ID, try to get channel name
                if channel.isdigit():
                    try:
                        # This will be processed later when we have ctx access
                        processed_channels.append(channel)
                    except:
                        processed_channels.append(channel)
                else:
                    processed_channels.append(channel)

            filters['channels'] = processed_channels
            logger.info(f"Channels added to filters: {filters['channels']}")
        
        # Extract time range
        time_match = re.search(r'(\d+)([dwmyh])', command_text)
        if time_match:
            amount, unit = time_match.groups()
            filters['time_range'] = {'amount': int(amount), 'unit': unit}

        return filters
    
    def apply_filters(self, data: Dict[str, Any], filters: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Apply all filters to user data - OPTIMIZED"""
        results = []
        
        for user_id, user_data in iter_user_entries(data):
            if self._user_matches_filters_optimized(user_data, filters):
                # Extract counter data
                counter = SmartCounter.from_dict(user_data.get('counter', {}))
                content_counter = ContentCounter.from_dict(user_data.get('content_counter', {}))
                
                results.append({
                    'user_id': user_id,
                    'username': user_data.get('username', 'Unknown'),
                    'roles': user_data.get('roles', []),
                    'total_messages': counter.total_count,
                    'counter': counter,
                    'content_counter': content_counter,
                    'first_seen': counter.first_seen,
                    'last_seen': counter.last_seen
                })
        
        # Sort by message count descending
        results.sort(key=lambda x: x['total_messages'], reverse=True)
        return results
    
    def _user_matches_filters_optimized(self, user_data: Dict[str, Any], filters: Dict[str, Any]) -> bool:
        """Check if user matches all filters - OPTIMIZED VERSION"""
        counter = SmartCounter.from_dict(user_data.get('counter', {}))
        content_counter = ContentCounter.from_dict(user_data.get('content_counter', {}))
        
        # Time range filter - check first
        time_range_days = None
        time_range_hours = None
        time_range_count = None

        if filters.get('time_range'):
            time_range_unit = filters['time_range']['unit']
            if time_range_unit == 'h':
                time_range_hours = self._convert_time_to_hours(filters['time_range'])
                time_range_count = counter.get_hourly_count(time_range_hours)
            else:
                time_range_days = self._convert_time_to_days(filters['time_range'])
                time_range_count = counter.get_time_range_count(time_range_days)

            # If user has no messages in time range, skip
            if time_range_count == 0:
                return False

        # Check message count filter (with or without time range)
        if filters.get('message_count'):
            op = filters['message_count']['operator']
            val = filters['message_count']['value']

            # Determine what count to check based on filters
            if filters.get('channels'):
                # Channel filter specified - use channel-specific count
                if time_range_hours is not None:
                    # Hour-based time range
                    channel_time_range_count = 0
                    for channel in filters['channels']:
                        channel_time_range_count += counter.get_channel_count_hours(channel, time_range_hours)
                    count_to_check = channel_time_range_count
                elif time_range_days is not None:
                    # With time range - use channel-specific time range count
                    channel_time_range_count = 0
                    for channel in filters['channels']:
                        channel_time_range_count += counter.get_channel_count(channel, time_range_days)
                    count_to_check = channel_time_range_count
                else:
                    # No time range - use total channel count
                    count_to_check = sum(counter.get_channel_count(channel) for channel in filters['channels'])
            else:
                # No channel filter - use total count
                if time_range_hours is not None:
                    count_to_check = time_range_count
                elif time_range_days is not None:
                    count_to_check = time_range_count
                else:
                    count_to_check = counter.total_count

            if op == '>' and not count_to_check > val:
                return False
            elif op == '<' and not count_to_check < val:
                return False
            elif op == '>=' and not count_to_check >= val:
                return False
            elif op == '<=' and not count_to_check <= val:
                return False
            elif op == '==' and not count_to_check == val:
                return False
            elif op == '!=' and not count_to_check != val:
                return False
        
        # Role filter
        if filters.get('roles', {}).get('include'):
            user_roles = [r.lower() for r in user_data.get('roles', [])]
            user_role_ids = user_data.get('role_ids', [])
            required_roles = [r.lower() for r in filters['roles']['include']]

            # Check for role ID matches (<@&123456789> format)
            role_id_matches = False
            role_id_required_roles = []
            regular_required_roles = []

            for role in required_roles:
                if role.startswith('<@&') and role.endswith('>'):
                    # Extract role ID from <@&123456789> format
                    role_id = role[3:-1]
                    role_id_required_roles.append(role_id)
                    # Check if user has this role ID
                    if role_id in user_role_ids:
                        role_id_matches = True
                else:
                    regular_required_roles.append(role)

            # For OR logic: if any role ID matches OR any regular role matches, user passes
            if filters['roles']['operator'] == 'or':
                if role_id_matches or any(role in user_roles for role in regular_required_roles):
                    pass  # User passes filter
                else:
                    return False

            # For AND logic: must match ALL role IDs AND ALL regular roles
            elif filters['roles']['operator'] == 'and':
                # Check all role ID requirements
                if role_id_required_roles:
                    if not all(role_id in user_role_ids for role_id in role_id_required_roles):
                        return False

                # Check all regular role requirements
                if regular_required_roles:
                    if not all(role in user_roles for role in regular_required_roles):
                        return False

        # Check exclude roles (nohave) - user must NOT have any of these roles
        if filters.get('roles', {}).get('exclude'):
            exclude_roles = [r.lower() for r in filters['roles']['exclude']]

            # Check for role ID exclusions
            for exclude_role in exclude_roles:
                if exclude_role.startswith('<@&') and exclude_role.endswith('>'):
                    # Extract role ID from <@&123456789> format
                    exclude_role_id = exclude_role[3:-1]
                    # If user has this excluded role ID, filter them out
                    if exclude_role_id in user_role_ids:
                        return False
                else:
                    # If user has this excluded role name, filter them out
                    if exclude_role in user_roles:
                        return False

        # Message count filter (if no time range specified)
        if filters.get('message_count') and not filters.get('time_range'):
            # If channel filter is specified, use channel-specific message count
            if filters.get('channels'):
                # Use the sum of messages from all specified channels
                msg_count = sum(counter.channel_counts.get(channel, 0) for channel in filters['channels'])
            else:
                msg_count = counter.total_count
            op = filters['message_count']['operator']
            val = filters['message_count']['value']
            
            if op == '>' and not msg_count > val:
                return False
            elif op == '<' and not msg_count < val:
                return False
            elif op == '>=' and not msg_count >= val:
                return False
            elif op == '<=' and not msg_count <= val:
                return False
            elif op == '==' and not msg_count == val:
                return False
            elif op == '!=' and not msg_count != val:
                return False
        
        # Content count filter
        if filters.get('content_count'):
            # If time range is specified, use time range content count
            if filters.get('time_range'):
                if time_range_hours is not None:
                    content_range_count = content_counter.get_time_range_count_hours(time_range_hours)
                else:
                    days = time_range_days if time_range_days is not None else self._convert_time_to_days(filters['time_range'])
                    content_range_count = content_counter.get_time_range_count(days)
                content_count = content_range_count
            else:
                # Use total content count
                content_count = len(content_counter.content_links)
            
            op = filters['content_count']['operator']
            val = filters['content_count']['value']
            
            if op == '>' and not content_count > val:
                return False
            elif op == '<' and not content_count < val:
                return False
            elif op == '>=' and not content_count >= val:
                return False
            elif op == '<=' and not content_count <= val:
                return False
            elif op == '==' and not content_count == val:
                return False
            elif op == '!=' and not content_count != val:
                return False

        # Channel filter - check if user has activity in specified channels within time range
        if filters.get('channels'):
            required_channels = set(filters['channels'])

            logger.info(f"Required channels: {required_channels}")

            # If time range is specified, check channel activity within that time range
            if filters.get('time_range'):
                has_channel_activity = False

                if time_range_hours is not None:
                    for channel in required_channels:
                        channel_count = counter.get_channel_count_hours(channel, time_range_hours)
                        if channel_count > 0:
                            has_channel_activity = True
                            logger.info(f"User has {channel_count} messages in {channel} within time range (hours)")
                            break
                else:
                    days = time_range_days if time_range_days is not None else self._convert_time_to_days(filters['time_range'])
                    for channel in required_channels:
                        channel_count = counter.get_channel_count(channel, days)
                        if channel_count > 0:
                            has_channel_activity = True
                            logger.info(f"User has {channel_count} messages in {channel} within time range")
                            break

                if not has_channel_activity:
                    logger.info(f"User filtered out - no activity in required channels within time range")
                    return False
                else:
                    logger.info(f"User passed channel filter with time range")
            else:
                # No time range - check if user has any messages in these channels
                user_channels = set(counter.channel_counts.keys())
                if not any(channel in user_channels for channel in required_channels):
                    logger.info(f"User filtered out - no matching channels")
                    return False
                else:
                    logger.info(f"User passed channel filter")

        return True
    
    def _convert_time_to_days(self, time_range: Dict[str, Any]) -> int:
        """Convert time range to days"""
        amount = time_range['amount']
        unit = time_range['unit']
        
        if unit == 'd':
            return amount
        elif unit == 'w':
            return amount * 7
        elif unit == 'm':
            return amount * 30
        elif unit == 'y':
            return amount * 365
        elif unit == 'h':
            return amount / 24

        return amount

    def _convert_time_to_hours(self, time_range: Dict[str, Any]) -> int:
        """Convert time range to hours"""
        amount = time_range['amount']
        unit = time_range['unit']

        if unit == 'h':
            return amount
        elif unit == 'd':
            return amount * 24
        elif unit == 'w':
            return amount * 7 * 24
        elif unit == 'm':
            return amount * 30 * 24
        elif unit == 'y':
            return amount * 365 * 24

        return amount * 24
    
    def export_to_excel(self, results: List[Dict[str, Any]], filters: Dict[str, Any]) -> str:
        """Export filtered results to Excel - WITH PROPER HYPERLINKS"""
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        # Create temporary file
        temp_file = tempfile.NamedTemporaryFile(mode='w', suffix='.xlsx', delete=False)
        temp_file.close()

        # Create workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Filter Results"

        # Define headers
        headers = [
            'User ID', 'Username', 'Total Messages', 'Time Range Messages',
            'Message Count 24h', 'Message Count 7d', 'Message Count 30d',
            'Message Count 60d', 'Message Count 90d', 'Message Count 120d',
            'Message Count 180d', 'Message Count 360d', 'Content Count 24h',
            'Content Count 7d', 'Content Count 30d', 'Content Count 60d',
            'Content Count 90d', 'Content Count 120d', 'Content Count 180d',
            'Content Count 360d', 'Total Content Links', 'Content Links',
            'Roles', 'First Seen', 'Last Seen', 'First Content Seen', 'Last Content Seen'
        ]

        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")

        # Prepare and write data
        for row_num, user in enumerate(results, 2):
            counter = user['counter']
            content_counter = user['content_counter']

            # Calculate time range count if specified
            time_range_count = counter.total_count
            if filters.get('time_range'):
                if filters['time_range']['unit'] == 'h':
                    hours = self._convert_time_to_hours(filters['time_range'])
                    time_range_count = counter.get_hourly_count(hours)
                else:
                    days = self._convert_time_to_days(filters['time_range'])
                    time_range_count = counter.get_time_range_count(days)

            # Message statistics
            if filters.get('channels'):
                channel_msg_24h = sum(counter.get_channel_count(channel, 1) for channel in filters['channels'])
                channel_msg_7d = sum(counter.get_channel_count(channel, 7) for channel in filters['channels'])
                channel_msg_30d = sum(counter.get_channel_count(channel, 30) for channel in filters['channels'])
                channel_msg_60d = sum(counter.get_channel_count(channel, 60) for channel in filters['channels'])
                channel_msg_90d = sum(counter.get_channel_count(channel, 90) for channel in filters['channels'])
                channel_msg_120d = sum(counter.get_channel_count(channel, 120) for channel in filters['channels'])
                channel_msg_180d = sum(counter.get_channel_count(channel, 180) for channel in filters['channels'])
                channel_msg_360d = sum(counter.get_channel_count(channel, 360) for channel in filters['channels'])
                total_channel_msgs = sum(counter.channel_counts.get(channel, 0) for channel in filters['channels'])
            else:
                channel_msg_24h = counter.get_time_range_count(1)
                channel_msg_7d = counter.get_time_range_count(7)
                channel_msg_30d = counter.get_time_range_count(30)
                channel_msg_60d = counter.get_time_range_count(60)
                channel_msg_90d = counter.get_time_range_count(90)
                channel_msg_120d = counter.get_time_range_count(120)
                channel_msg_180d = counter.get_time_range_count(180)
                channel_msg_360d = counter.get_time_range_count(360)
                total_channel_msgs = counter.total_count

            # Content statistics
            content_count_24h = content_counter.get_time_range_count(1)
            content_count_7d = content_counter.get_time_range_count(7)
            content_count_30d = content_counter.get_time_range_count(30)
            content_count_60d = content_counter.get_time_range_count(60)
            content_count_90d = content_counter.get_time_range_count(90)
            content_count_120d = content_counter.get_time_range_count(120)
            content_count_180d = content_counter.get_time_range_count(180)
            content_count_360d = content_counter.get_time_range_count(360)
            total_content = len(content_counter.content_links)

            # Write basic data
            row_data = [
                user['user_id'],
                user['username'],
                total_channel_msgs,
                time_range_count,
                channel_msg_24h,
                channel_msg_7d,
                channel_msg_30d,
                channel_msg_60d,
                channel_msg_90d,
                channel_msg_120d,
                channel_msg_180d,
                channel_msg_360d,
                content_count_24h,
                content_count_7d,
                content_count_30d,
                content_count_60d,
                content_count_90d,
                content_count_120d,
                content_count_180d,
                content_count_360d,
                total_content,
                '',  # Content Links - will be filled with hyperlinks
                ', '.join(user['roles']),
                counter.first_seen[:19].replace('T', ' ') if counter.first_seen else 'N/A',
                counter.last_seen[:19].replace('T', ' ') if counter.last_seen else 'N/A',
                content_counter.first_content_seen[:19].replace('T', ' ') if content_counter.first_content_seen else 'N/A',
                content_counter.last_content_seen[:19].replace('T', ' ') if content_counter.last_content_seen else 'N/A',
            ]

            # Write row data
            for col_num, value in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col_num, value=value)

            # Add hyperlinks for content links - EACH LINK SEPARATELY CLICKABLE
            content_links_col = 22  # Content Links column
            if content_counter.content_links:
                # Create single hyperlink pointing to last link
                if content_counter.content_links:
                    # Combine all link URLs with newlines
                    combined_display = '\n'.join([link_data['link'] for link_data in content_counter.content_links])

                    # Set row height to accommodate multiple lines (each link ~20 pixels)
                    ws.row_dimensions[row_num].height = max(25, len(content_counter.content_links) * 20)

                    # Create cell with all links displayed
                    cell = ws.cell(row=row_num, column=content_links_col, value=combined_display)

                    # Set hyperlink to the LAST link (entire display text points to final link)
                    last_link = content_counter.content_links[-1]['link']
                    cell.hyperlink = last_link

                    # Style as hyperlink
                    cell.font = Font(color="0000FF", underline="single")
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
                else:
                    # No content links - show "No content links"
                    cell = ws.cell(row=row_num, column=content_links_col, value="No content links")
                    cell.alignment = Alignment(wrap_text=True, vertical='top')

        # Auto-fit column widths
        for col_num, header in enumerate(headers, 1):
            column_letter = get_column_letter(col_num)
            max_length = len(header) + 2

            # Check data length in this column
            for row_num in range(2, len(results) + 2):
                cell_value = ws.cell(row=row_num, column=col_num).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)) + 2)

            ws.column_dimensions[column_letter].width = min(max_length, 50)  # Cap at 50 characters

        # Save workbook
        wb.save(temp_file.name)

        return temp_file.name

# Initialize filter engine
filter_engine = FilterEngine()

def _load_data_from_disk():
    """Load message data from JSON file"""
    try:
        if DATA_FILE.exists():
            with open(DATA_FILE, 'r', encoding='utf-8') as f:
                data = json.load(f)

            # Migrate old format data to new format
            migrated = False
            for user_id, user_data in data.items():
                if _is_meta_key(user_id):
                    continue
                if 'messages' in user_data and 'counter' not in user_data:
                    data[user_id] = migrate_user_data(user_data)
                    migrated = True

            if migrated:
                logger.info("Data migration completed, saving migrated data")
                _save_data_to_disk(data)

            return data
        return {}
    except Exception as e:
        logger.error(f"Error loading data: {e}")
        return {}

def _save_data_to_disk(data):
    """Save message data to JSON file atomically"""
    try:
        # Write to temporary file first
        temp_file = DATA_FILE.with_suffix('.tmp')
        with open(temp_file, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
        
        # Atomic rename
        temp_file.rename(DATA_FILE)
        logger.info(f"Data saved successfully. Total users: {len(data)}")
        return True
    except Exception as e:
        logger.error(f"Error saving data: {e}")
        return False

async def ensure_cache_loaded() -> None:
    """Load message data into the in-memory cache if necessary"""
    global message_cache

    async with get_message_cache_lock():
        if message_cache:
            return

    loop = asyncio.get_running_loop()
    data = await loop.run_in_executor(None, _load_data_from_disk)

    async with get_message_cache_lock():
        if not message_cache:
            message_cache.clear()
            message_cache.update(data)

async def get_cached_data() -> Dict[str, Any]:
    """Get a deep copy of the cached message data"""
    await ensure_cache_loaded()
    return await get_message_cache_snapshot()

async def load_user_data(user_id):
    """Load specific user data from the cached data"""
    data = await get_cached_data()
    return data.get(user_id)

# update_user_data function removed - now handled by async processing

@bot.event
async def on_ready():
    """Bot ready event with content detection system"""
    logger.info(f'Bot logged in as {bot.user}!')
    logger.info('Starting content detection message tracker...')

    # Start file manager
    await file_manager.start()

    # Load content channels
    load_content_channels()
    logger.info(f"Loaded {len(content_channels)} content channels")

    # Log tracking configuration
    if TRACKED_ROLE_IDS:
        logger.info(f"Role-based tracking enabled - Tracking roles: {', '.join(TRACKED_ROLE_IDS)}")
    else:
        logger.info("No specific roles configured - tracking all users")

    # Load existing data into cache
    existing_data = await get_cached_data()
    user_count = sum(1 for _ in iter_user_entries(existing_data))
    logger.info(f"Loaded existing data: {user_count} users")

    # Migrate user role IDs if needed
    await migrate_user_role_ids()

    logger.info("Content detection message tracker started!")

@bot.event
async def on_disconnect():
    """Handle bot disconnect"""
    logger.info("Bot disconnecting, stopping file manager...")
    await file_manager.stop()

@bot.event
async def on_close():
    """Handle bot close"""
    logger.info("Bot closing, stopping file manager...")
    await file_manager.stop()

@async_task
async def _handle_member_update_async(user_id: str, username: str, roles: List[str], role_ids: List[str] = None):
    """Handle member update asynchronously - JSON DIRECT VERSION"""
    try:
        await update_user_in_json(user_id, username, roles, role_ids)
        logger.debug(f"Updated roles for user {user_id}: {roles} (IDs: {role_ids})")
    except Exception as e:
        logger.error(f"Error in async member update: {e}")

@bot.event
async def on_member_update(before, after):
    """Handle member role updates - JSON DIRECT VERSION"""
    try:
        # Check if roles or username changed
        if before.roles != after.roles or before.display_name != after.display_name:
            # Only track users who should be tracked
            if should_track_user(after):
                user_id = str(after.id)
                username = after.display_name
                roles = [role.name for role in after.roles if role.name != '@everyone']
                role_ids = [str(role.id) for role in after.roles if role.name != '@everyone']

                # Handle asynchronously
                _handle_member_update_async(user_id, username, roles, role_ids)

                logger.info(f"Updated user info for {user_id}: {username} (roles changed)")
            else:
                logger.debug(f"Skipping member update for {after.display_name} - not in tracked roles")

    except Exception as e:
        logger.error(f"Error in member update: {e}")

@bot.event
async def on_member_join(member):
    """Handle new member joining - JSON DIRECT VERSION"""
    try:
        # Only track users who should be tracked
        if should_track_user(member):
            user_id = str(member.id)
            username = member.display_name
            roles = [role.name for role in member.roles if role.name != '@everyone']
            role_ids = [str(role.id) for role in member.roles if role.name != '@everyone']

            # Handle asynchronously
            _handle_member_update_async(user_id, username, roles, role_ids)

            logger.info(f"Created user data for new member {user_id}: {username}")
        else:
            logger.debug(f"Skipping new member {member.display_name} - not in tracked roles")

    except Exception as e:
        logger.error(f"Error in member join: {e}")

@async_task
async def _process_message_async(message):
    """Process message asynchronously - JSON DIRECT VERSION"""
    try:
        user_id = str(message.author.id)

        # Create message info
        message_info = {
            'timestamp': datetime.now().isoformat(),
            'channel_name': message.channel.name
        }

        await ensure_cache_loaded()
        async with get_message_cache_lock():
            user_entry = message_cache.get(user_id)

            if user_entry:
                counter = SmartCounter.from_dict(user_entry.get('counter', {}))
            else:
                counter = SmartCounter()
                user_entry = {
                    'username': message.author.display_name,
                    'roles': [role.name for role in message.author.roles if role.name != '@everyone'],
                    'counter': counter.to_dict(),
                    'role_ids': [str(role.id) for role in message.author.roles if role.name != '@everyone']
                }
                message_cache[user_id] = user_entry

            counter.add_message(message_info['timestamp'], message_info['channel_name'])
            user_entry['counter'] = counter.to_dict()

        file_manager.mark_dirty()

        # Update stats
        tracker_stats['messages_processed'] += 1

        # Log performance
        if tracker_stats['messages_processed'] % 5000 == 0:
            elapsed = time.time() - tracker_stats['start_time']
            msg_per_sec = tracker_stats['messages_processed'] / elapsed if elapsed > 0 else 0
            api_calls_per_msg = tracker_stats['api_calls'] / max(tracker_stats['messages_processed'], 1)
            logger.info(f"🚀 JSON DIRECT Performance: {tracker_stats['messages_processed']} messages, "
                       f"{msg_per_sec:.1f} msg/s, API/msg: {api_calls_per_msg:.3f}, "
                       f"Writes: {tracker_stats['disk_writes']}")
    
    except Exception as e:
        logger.error(f"Error in async message processing: {e}")
        logger.error(f"Message type: {type(message)}")
        logger.error(f"Message author type: {type(message.author)}")
        logger.error(f"Message details: {message}")
        import traceback
        logger.error(f"Full traceback: {traceback.format_exc()}")

@bot.event
async def on_message(message):
    """Handle incoming messages with content detection"""
    # Skip bot messages
    if message.author.bot:
        return

    try:
        if message.content and message.content.startswith('!'):
            if not has_full_access(message.author) and not is_authorized_member(message.author):
                try:
                    await message.delete()
                except Exception as delete_error:
                    logger.debug(f'Failed to delete unauthorized command: {delete_error}')
                return

        # Check if user should be tracked
        if should_track_user(message.author):
            # Process message asynchronously
            _process_message_async(message)

            # Check for content links if this is a content channel
            channel_id = str(message.channel.id)
            is_content = is_content_channel(channel_id)
            logger.info(f"DEBUG: Channel {channel_id} ({message.channel.name}) is content channel: {is_content}")

            if is_content:
                timestamp = datetime.now().isoformat()
                detected_links, new_links = await process_content_links(
                    str(message.author.id),
                    message.content,
                    timestamp
                )

                if detected_links:
                    logger.info(
                        "Detected %s content links from %s (new: %s)",
                        len(detected_links),
                        message.author.name,
                        len(new_links)
                    )
        else:
            logger.debug(f"Skipping message from {message.author.name} - not in tracked roles")

        if message.content and message.content.startswith('!'):
            await bot.process_commands(message)

    except Exception as e:
        logger.error(f"Error processing message: {e}")

@bot.command()
async def stats(ctx):
    """Show bot statistics with JSON DIRECT system"""
    try:
        elapsed = time.time() - tracker_stats['start_time']
        msg_per_sec = tracker_stats['messages_processed'] / elapsed if elapsed > 0 else 0

        existing_data = await get_cached_data()
        total_users = sum(1 for _ in iter_user_entries(existing_data))

        # Calculate totals from counters
        total_messages = 0
        all_channels = set()
        all_roles = set()
        last_24_hours = 0
        last_7_days = 0
        last_30_days = 0

        for _, user_data in iter_user_entries(existing_data):
            counter = SmartCounter.from_dict(user_data.get('counter', {}))
            total_messages += counter.total_count

            # Collect channels and roles
            all_channels.update(counter.channel_counts.keys())
            all_roles.update(user_data.get('roles', []))

            # Calculate time ranges with accurate 24-hour counting
            last_24_hours += counter.get_time_range_count(1)
            last_7_days += counter.get_time_range_count(7)
            last_30_days += counter.get_time_range_count(30)

        # Calculate API metrics
        api_calls_per_msg = tracker_stats['api_calls'] / max(tracker_stats['messages_processed'], 1)
        api_savings = max(tracker_stats['messages_processed'] - tracker_stats['api_calls'], 0)
        api_efficiency = ((api_savings / max(tracker_stats['messages_processed'], 1)) * 100) if tracker_stats['messages_processed'] > 0 else 0

        pending_status = 'Yes' if file_manager.dirty else 'No'
        storage_rate = tracker_stats['disk_writes'] / elapsed if elapsed > 0 else 0

        lines = [
            "🚀 Bot Statistics (JSON DIRECT)",
            f"💬 Messages: processed {tracker_stats['messages_processed']:,} | rate {msg_per_sec:.1f}/s",
            f"👥 Users: total {total_users:,}" if total_users > 0 else "👥 Users: total 0",
            f"📊 Data: total {total_messages:,} | 24h {last_24_hours:,} | 7d {last_7_days:,} | 30d {last_30_days:,}",
            f"📺 Channels tracked: {len(all_channels)}",
            f"👑 Roles recorded: {len(all_roles)}",
            f"⚡ API calls: {tracker_stats['api_calls']:,} | per msg {api_calls_per_msg:.3f}",
            f"💾 API savings: {api_savings:,} | efficiency {api_efficiency:.1f}%",
            f"🛡️ Rate limits hit: {tracker_stats['rate_limit_hits']:,}",
            f"💽 Storage: pending flush {pending_status} | write rate {storage_rate:.2f}/s",
            f"⏱️ Uptime: {elapsed/3600:.1f} hours"
        ]

        if total_users > 0:
            avg_messages = total_messages / total_users if total_users else 0
            avg_roles = len(all_roles) / total_users if total_users else 0
            lines.insert(3, f"📈 Averages: messages/user {avg_messages:.1f} | roles/user {avg_roles:.1f}")

        await ctx.send("\n".join(lines))
            
    except Exception as e:
        await ctx.send(f"Error getting stats: {e}")

@bot.command()
async def api_stats(ctx):
    """Show detailed API usage statistics"""
    try:
        elapsed = time.time() - tracker_stats['start_time']

        # Calculate detailed metrics
        api_calls_per_minute = (tracker_stats['api_calls'] / elapsed) * 60 if elapsed > 0 else 0
        efficiency = ((tracker_stats['messages_processed'] - tracker_stats['api_calls']) / tracker_stats['messages_processed']) * 100 if tracker_stats['messages_processed'] > 0 else 0

        lines = [
            "🔧 Detailed API Statistics",
            f"📈 API usage: total calls {tracker_stats['api_calls']:,} | per minute {api_calls_per_minute:.1f}",
            "🎯 Cache efficiency: hit rate N/A (direct JSON)",
            f"💰 Savings: efficiency {efficiency:.1f}% | saved calls {tracker_stats['messages_processed'] - tracker_stats['api_calls']:,}",
            f"⚠️ Rate limits: hits {tracker_stats['rate_limit_hits']:,} | rate {tracker_stats['rate_limit_hits']/elapsed:.2f}/s" if elapsed > 0 else "⚠️ Rate limits: hits 0",
            f"💾 Disk writes: total {tracker_stats['disk_writes']:,} | rate {tracker_stats['disk_writes']/elapsed:.2f}/s" if elapsed > 0 else "💾 Disk writes: total 0",
            "🗂️ Cache size: N/A (direct JSON access)"
        ]

        await ctx.send("\n".join(lines))
            
    except Exception as e:
        await ctx.send(f"Error getting API stats: {e}")

@bot.command()
async def save_now(ctx):
    """Manual save command - flush write queue"""
    try:
        if file_manager.dirty:
            await ctx.send("📝 Flushing pending writes...")
            await file_manager.flush_now()
            await ctx.send("✅ All pending writes completed!")
        else:
            await ctx.send("✅ No pending writes - all data saved.")
            
    except Exception as e:
        await ctx.send(f"Error processing save: {e}")

@bot.command()
async def addcontentchannel(ctx, channel: discord.TextChannel = None):
    """Add a channel for content link detection"""
    try:
        if channel is None:
            channel = ctx.channel
        
        channel_id = str(channel.id)
        
        if channel_id in content_channels:
            await ctx.send(f"❌ {channel.mention} is already a content channel!")
            return
        
        content_channels.add(channel_id)
        await save_content_channels()
        
        await ctx.send(f"✅ {channel.mention} added as content channel!\n"
                   f"Channel ID: {channel_id}\n"
                   f"Total content channels: {len(content_channels)}\n"
                   f"Now monitoring {channel.mention} for X (Twitter) links!")
        logger.info(f"Added content channel: {channel.name} ({channel_id})")
        
    except Exception as e:
        await ctx.send(f"Error adding content channel: {e}")

@bot.command()
async def removecontentchannel(ctx, channel: discord.TextChannel = None):
    """Remove a channel from content link detection"""
    try:
        if channel is None:
            channel = ctx.channel
        
        channel_id = str(channel.id)
        
        if channel_id not in content_channels:
            await ctx.send(f"❌ {channel.mention} is not a content channel!")
            return
        
        content_channels.remove(channel_id)
        await save_content_channels()
        
        await ctx.send(f"➖ {channel.mention} removed from content channels!\n"
                   f"Channel ID: {channel_id}\n"
                   f"Total content channels: {len(content_channels)}\n"
                   f"No longer monitoring {channel.mention} for content links.")
        logger.info(f"Removed content channel: {channel.name} ({channel_id})")
        
    except Exception as e:
        await ctx.send(f"Error removing content channel: {e}")

@bot.command()
async def listcontentchannels(ctx):
    """List all content channels"""
    try:
        if not content_channels:
            await ctx.send("❌ No content channels configured!")
            return
        
        channel_list = []
        for channel_id in sorted(content_channels):
            try:
                channel = bot.get_channel(int(channel_id))
                if channel:
                    channel_list.append(f"- {channel.mention} ({channel_id})")
                else:
                    channel_list.append(f"- Unknown channel ({channel_id})")
            except:
                channel_list.append(f"- Invalid channel ({channel_id})")

        header = f"📋 Content Channels ({len(content_channels)} total)"
        messages = []
        current_lines = [header]
        current_length = len(header)

        for line in channel_list:
            line_length = len(line) + 1
            if current_length + line_length > 1900:
                messages.append("\n".join(current_lines))
                current_lines = ["📋 Content Channels (continued)"]
                current_length = len(current_lines[0])
            current_lines.append(line)
            current_length += line_length

        if current_lines:
            messages.append("\n".join(current_lines))

        for message in messages:
            await ctx.send(message)
        
    except Exception as e:
        await ctx.send(f"Error listing content channels: {e}")

@bot.command()
async def addauthorized(ctx, entry_type: str, *identifiers: str):
    """Add authorized users or roles (security manager only)"""
    if not is_security_manager(ctx.author):
        await ctx.send('❌ Only the security manager can add authorized entries.')
        return

    if not identifiers:
        await ctx.send('Usage: `!addauthorized <user|role> <id or mention> [id ...]`')
        return

    entry = entry_type.lower().strip()
    if entry not in {'user', 'role'}:
        await ctx.send('Usage: `!addauthorized <user|role> <id or mention> [id ...]`')
        return

    added = []
    skipped = []

    for raw_identifier in identifiers:
        success, normalized, resolved_type = await add_authorized_identifier(entry, raw_identifier)
        if success and normalized:
            label = (resolved_type or entry).lower()
            added.append((label, normalized))
            logger.info(f'Added authorized {label}: {normalized}')
        else:
            if normalized:
                label = (resolved_type or entry).lower() if resolved_type else entry
                skipped.append((label, normalized))
            else:
                skipped.append((entry, raw_identifier))

    if added:
        lines = '\n'.join(f'✅ {label}: `{identifier}`' for label, identifier in added)
        await ctx.send(f'Added {len(added)} authorized entries:\n{lines}')
    else:
        await ctx.send('No entries were added.')

    if skipped:
        lines = '\n'.join(f'• {label}: `{identifier}`' for label, identifier in skipped)
        await ctx.send(f'Skipped entries (already authorized or invalid):\n{lines}')


@bot.command()
async def removeauthorized(ctx, entry_type: str, *identifiers: str):
    """Remove authorized users or roles (security manager only)"""
    if not is_security_manager(ctx.author):
        await ctx.send('❌ Only the security manager can modify authorized entries.')
        return

    if not identifiers:
        await ctx.send('Usage: `!removeauthorized <user|role> <id or mention> [id ...]`')
        return

    entry = entry_type.lower().strip()
    if entry not in {'user', 'role'}:
        await ctx.send('Usage: `!removeauthorized <user|role> <id or mention> [id ...]`')
        return

    removed = []
    skipped = []

    for raw_identifier in identifiers:
        success, normalized, resolved_type = await remove_authorized_identifier(entry, raw_identifier)
        if success and normalized:
            label = (resolved_type or entry).lower()
            removed.append((label, normalized))
            logger.info(f'Removed authorized {label}: {normalized}')
        else:
            if normalized:
                label = (resolved_type or entry).lower() if resolved_type else entry
                skipped.append((label, normalized))
            else:
                skipped.append((entry, raw_identifier))

    if removed:
        lines = '\n'.join(f'🗑️ {label}: `{identifier}`' for label, identifier in removed)
        await ctx.send(f'Removed {len(removed)} authorized entries:\n{lines}')
    else:
        await ctx.send('No entries were removed.')

    if skipped:
        lines = '\n'.join(f'• {label}: `{identifier}`' for label, identifier in skipped)
        await ctx.send(f'Skipped entries (not found or invalid):\n{lines}')


@bot.command()
async def listauthorized(ctx):
    """List all authorized users and roles"""
    if not has_full_access(ctx.author):
        await ctx.send('❌ You do not have permission to view authorized entries.')
        return

    users, roles = await get_authorization_snapshot()

    lines = [
        '🔐 Authorized Access',
        'Users:' if users else 'Users: none'
    ]

    if users:
        lines.extend(f'- {user_id}' for user_id in users)

    if roles:
        lines.append('Roles:')
        lines.extend(f'- {role_id}' for role_id in roles)
    else:
        lines.append('Roles: none')

    await ctx.send("\n".join(lines))



@bot.command()
async def system_info(ctx):
    """Show system information"""
    try:
        existing_data = await get_cached_data()
        total_users = sum(1 for _ in iter_user_entries(existing_data))
        
        # Calculate total messages
        total_messages = 0
        total_content_links = 0
        for _, user_data in iter_user_entries(existing_data):
            counter = SmartCounter.from_dict(user_data.get('counter', {}))
            total_messages += counter.total_count
            
            # Count content links
            if 'content_counter' in user_data:
                content_counter = ContentCounter.from_dict(user_data['content_counter'])
                total_content_links += content_counter.get_total_content_count()
        
        elapsed = time.time() - tracker_stats['start_time']
        msg_per_sec = tracker_stats['messages_processed'] / elapsed if elapsed > 0 else 0
        api_calls_per_msg = tracker_stats['api_calls'] / max(tracker_stats['messages_processed'], 1)
        
        pending_status = 'Yes' if file_manager.dirty else 'No'

        lines = [
            "🖥️ System Information",
            f"📊 Users: {total_users:,} | Messages: {total_messages:,}",
            f"🔗 Content links: total {total_content_links:,} | channels tracked {len(content_channels)}",
            f"⚡ Performance: processed {tracker_stats['messages_processed']:,} | rate {msg_per_sec:.1f}/s",
            f"🔧 API usage: total {tracker_stats['api_calls']:,} | per message {api_calls_per_msg:.3f}",
            f"💾 Storage: pending flush {pending_status} | writes {tracker_stats['disk_writes']:,}",
            f"⏱️ Uptime: {elapsed/3600:.1f} hours"
        ]

        await ctx.send("\n".join(lines))
            
    except Exception as e:
        await ctx.send(f"Error getting system info: {e}")

@bot.command()
async def user_info(ctx, user_id: str = None):
    """Get information about a specific user with content stats"""
    try:
        if user_id is None:
            user_id = str(ctx.author.id)
        
        # Check saved data
        existing_data = await get_cached_data()
        user_data = existing_data.get(user_id)
        
        if not user_data:
            await ctx.send("User not found in database.")
            return
        
        # Get counter data
        counter = SmartCounter.from_dict(user_data.get('counter', {}))
        username = user_data.get('username', 'Unknown')
        roles = user_data.get('roles', [])
        
        lines = [
            f"🧾 User Info: {username}",
            f"ID: {user_id}",
            f"Total messages: {counter.total_count:,}"
        ]

        # Time range statistics with accurate 24-hour counting
        last_24h = counter.get_time_range_count(1)  # Uses the new 24-hour rolling window
        last_7d = counter.get_time_range_count(7)
        last_30d = counter.get_time_range_count(30)
        lines.append(f"Recent activity: 24h {last_24h}, 7d {last_7d}, 30d {last_30d}")

        # Content statistics
        content_summary = "No content links"
        latest_link_line = None
        if 'content_counter' in user_data:
            content_counter = ContentCounter.from_dict(user_data['content_counter'])
            total_content = content_counter.get_total_content_count()
            content_24h = content_counter.get_time_range_count(1)  # 24 hours = 1 day
            content_7d = content_counter.get_time_range_count(7)
            content_30d = content_counter.get_time_range_count(30)
            
            if total_content > 0:
                content_summary = (
                    f"Content links: total {total_content}, 24h {content_24h}, 7d {content_7d}, 30d {content_30d}"
                )
                
                # Show latest link if available
                if content_counter.content_links:
                    latest_link = content_counter.content_links[-1]['link']
                    if len(latest_link) > 50:
                        latest_link = latest_link[:47] + "..."
                    latest_link_line = f"Latest content link: {latest_link}"

        lines.append(content_summary)
        if latest_link_line:
            lines.append(latest_link_line)
        
        # Timestamps
        if counter.first_seen:
            first_time = datetime.fromisoformat(counter.first_seen.replace('Z', '+00:00')).strftime('%Y-%m-%d %H:%M')
            lines.append(f"First seen: {first_time}")
        
        if counter.last_seen:
            last_time = datetime.fromisoformat(counter.last_seen.replace('Z', '+00:00')).strftime('%Y-%m-%d %H:%M')
            lines.append(f"Last seen: {last_time}")
        
        # Channel statistics
        if counter.channel_counts:
            top_channels = sorted(counter.channel_counts.items(), key=lambda x: x[1], reverse=True)[:3]
            most_active_channel = top_channels[0] if top_channels else ('None', 0)
            
            lines.append(f"Top channels (most active: {most_active_channel[0]}):")
            lines.extend(f"- {chan}: {count:,}" for chan, count in top_channels)
        
        lines.append("Roles: " + (", ".join(roles) if roles else "None"))

        await ctx.send("\n".join(lines))

    except Exception as e:
        await ctx.send(f"Error getting user info: {e}")

@bot.command()
async def top_users(ctx, limit: int = 10):
    """Show top users by message count with content stats"""
    try:
        existing_data = await get_cached_data()
        
        # Prepare users with counts
        user_counts = []
        for user_id, user_data in iter_user_entries(existing_data):
            counter = SmartCounter.from_dict(user_data.get('counter', {}))
            
            # Use total count
            message_count = counter.total_count
            
            # Get content stats
            content_count = 0
            if 'content_counter' in user_data:
                content_counter = ContentCounter.from_dict(user_data['content_counter'])
                content_count = content_counter.get_total_content_count()
            
            user_counts.append({
                'user_id': user_id,
                'username': user_data.get('username', 'Unknown'),
                'message_count': message_count,
                'content_count': content_count,
                'counter': counter
            })
        
        # Sort by message count
        user_counts.sort(key=lambda x: x['message_count'], reverse=True)
        
        header = f"🏆 Top {limit} Users by Message Count"
        lines = [header]

        for i, user_info in enumerate(user_counts[:limit], 1):
            username = user_info['username']
            message_count = user_info['message_count']
            content_count = user_info['content_count']
            counter = user_info['counter']

            if counter.channel_counts:
                top_channel = max(counter.channel_counts.items(), key=lambda x: x[1])
                channel_text = f"{top_channel[0]} ({top_channel[1]:,})"
            else:
                channel_text = "No channels"

            last_24h = counter.get_time_range_count(1)
            last_7d = counter.get_time_range_count(7)
            content_label = f"Content links: {content_count}" if content_count > 0 else "Content links: none"

            lines.extend([
                f"{i}. {username} – {message_count:,} messages",
                f"   Activity: 24h {last_24h:,}, 7d {last_7d:,}",
                f"   {content_label}",
                f"   Top channel: {channel_text}"
            ])

        # Split into messages below Discord limit
        messages = []
        current = []
        current_len = 0
        for line in lines:
            line_len = len(line) + 1
            if current and current_len + line_len > 1900:
                messages.append("\n".join(current))
                current = ["(cont.)"]
                current_len = len(current[0])
            current.append(line)
            current_len += line_len

        if current:
            messages.append("\n".join(current))

        for message in messages:
            await ctx.send(message)
        
    except Exception as e:
        await ctx.send(f"Error getting top users: {e}")

@bot.command()
async def trackcommands(ctx):
    """Display all available commands with examples"""
    help_text = """
🤖 **Tracker Bot Commands - 24-Hour Counting Updated!**

**Content Management:**
• `!addcontentchannel #channel` - Add a content tracking channel
• `!removecontentchannel #channel` - Remove a content tracking channel
• `!listcontentchannels` - List tracked content channels

**User Information:**
• `!user_info @user` - Detailed user statistics with 24h data
• `!top_users [amount]` - Top users by message count
• `!stats` - Bot statistics with accurate 24h counting

**Advanced Filtering:**
• `!filter [filters]` - Advanced filtering with Excel export
• `!system_info` - Bot system information

**Security Management (Security Manager only):**
• `!addauthorized user <id>` - Add an authorized user
• `!addauthorized role <id>` - Add an authorized role
• `!removeauthorized user <id>` - Remove an authorized user
• `!removeauthorized role <id>` - Remove an authorized role
• `!listauthorized` - List authorized users and roles

**NEW FEATURE:**
• **Precise 24-Hour Count:** Message totals now use a true 24-hour rolling window
• **Real-Time:** Counts always reflect the last 24 hours accurately
• **Auto Reset:** Day rollover issues are fully resolved

**Filter Syntax Guide:**
**Message Filters:** `message>10`, `message<5`, `message>=20`
**Content Filters:** `content>2`, `content<10`, `content==5`
**Role Filters:** `@role1 and @role2`, `@role1 or @role2`
**Time Filters:** `7d`, `1w`, `30d`, `1m`, `1y`
**Operators:** `>`, `<`, `>=`, `<=`, `==`, `!=`

**Examples:**
• `!filter message>10 content>2 @role1 and @role2 7d`
• `!filter @admin or @moderator content>5 30d`
• `!filter message>15 content>3 @early agi and @advanced agi 7d`

**Excel Export Features:**
Generated Excel includes:
• User ID and Username
• Total Messages & Content Links
• **Time-based statistics (24h, 7d, 30d, 60d, 90d, 120d, 180d, 360d)**
• Role information
• First/Last seen timestamps
• Channel activity data
• Content sharing statistics

**Note:** Excel export is generated automatically when you run `!filter`
"""
    await ctx.send(help_text)

@bot.command()
async def helpme(ctx):
    """Simple help command without embed"""
    help_text = """
🤖 **Tracker Bot Commands:**

**Core Commands:**
• `!stats` - Show bot statistics
• `!trackcommands` - List all commands (detailed)
• `!system_info` - System information

**User Insights:**
• `!user_info @user` - User details
• `!top_users [count]` - Most active users

**Filtering:**
• `!filter message>10 @role1 and @role2 7d` - Advanced filtering
• `!help_filter` - Filtering help

**Content Tracking:**
• `!addcontentchannel #channel` - Add a content channel
• `!listcontentchannels` - List content channels

**Security Management (Security Manager only):**
• `!addauthorized user <id>` - Add an authorized user
• `!addauthorized role <id>` - Add an authorized role
• `!removeauthorized user <id>` - Remove an authorized user
• `!removeauthorized role <id>` - Remove an authorized role
• `!listauthorized` - List authorized entries

**NEW:** 24-hour message counting is now fully accurate!
**Note:** The bot currently tracks only users with role ID `1346571303657930804`.
"""
    await ctx.send(help_text)

@bot.command()
async def test_24h(ctx):
    """Test 24-hour message counting accuracy"""
    try:
        # Get current user's data
        user_id = str(ctx.author.id)
        existing_data = await get_cached_data()
        user_data = existing_data.get(user_id)

        if not user_data:
            await ctx.send("❌ User data not found. Please send some messages first.")
            return

        counter = SmartCounter.from_dict(user_data.get('counter', {}))

        # Calculate different time ranges
        total_24h = counter.get_time_range_count(1)
        total_7d = counter.get_time_range_count(7)
        total_30d = counter.get_time_range_count(30)

        # Get timestamp info
        now = datetime.now()
        cutoff_24h = now - timedelta(hours=24)

        # Manual count for verification
        if counter.message_timestamps:
            manual_24h = sum(1 for ts in counter.message_timestamps
                           if datetime.fromisoformat(ts.replace('Z', '+00:00')) >= cutoff_24h)
        else:
            manual_24h = 0

        lines = [
            "🧪 24-Hour Count Test",
            f"User: {ctx.author.display_name}",
            f"📊 Total messages: {counter.total_count:,}",
            f"   24h (system): {total_24h:,}",
            f"   24h (manual): {manual_24h:,}",
            f"   7d: {total_7d:,}",
            f"   30d: {total_30d:,}",
            f"🕐 Now: {now.strftime('%Y-%m-%d %H:%M:%S')}",
            f"   24h ago: {cutoff_24h.strftime('%Y-%m-%d %H:%M:%S')}",
            f"   Stored timestamps: {len(counter.message_timestamps)}"
        ]

        if total_24h == manual_24h:
            lines.append("✅ Verification: 24-hour counting is working correctly!")
        else:
            diff = abs(total_24h - manual_24h)
            lines.append(f"⚠️ Mismatch: system {total_24h}, manual {manual_24h} (difference {diff})")

        await ctx.send("\n".join(lines))

    except Exception as e:
        await ctx.send(f"❌ Error during test: {e}")
        logger.error(f"Error in test_24h: {e}")

@bot.command()
async def test_channel_filter(ctx, channel_name: str = None, min_messages: int = 25):
    """Test channel-specific 24-hour filtering"""
    try:
        if not channel_name:
            await ctx.send("❌ Please provide a channel name. Example: `!test_channel_filter #general 25`")
            return

        # Clean channel name (remove # if present)
        if channel_name.startswith('#'):
            channel_name = channel_name[1:]

        # Get current user's data
        user_id = str(ctx.author.id)
        existing_data = await get_cached_data()
        user_data = existing_data.get(user_id)

        if not user_data:
            await ctx.send("❌ User data not found. Please send some messages first.")
            return

        counter = SmartCounter.from_dict(user_data.get('counter', {}))

        # Test different scenarios
        # 1. Total messages in channel
        total_channel = counter.get_channel_count(channel_name)
        # 2. Messages in last 24 hours in channel
        channel_24h = counter.get_channel_count(channel_name, 1)
        # 3. Messages in last 7 days in channel
        channel_7d = counter.get_channel_count(channel_name, 7)

        filter_24h_pass = channel_24h >= min_messages
        filter_7d_pass = channel_7d >= min_messages

        lines = [
            f"🧪 Channel Filter Test: #{channel_name}",
            f"User: {ctx.author.display_name}",
            f"📊 Total messages: {total_channel:,}",
            f"   24h: {channel_24h:,}",
            f"   7d: {channel_7d:,}",
            f"🔍 Threshold: {min_messages} messages",
            f"   24h >= threshold: {'✅' if filter_24h_pass else '❌'}",
            f"   7d >= threshold: {'✅' if filter_7d_pass else '❌'}"
        ]

        if channel_name in counter.channel_timestamps:
            timestamps = counter.channel_timestamps[channel_name]
            lines.append(f"⏱️ Stored timestamps: {len(timestamps)}")
            if timestamps:
                last_timestamps = sorted(timestamps)[-3:]
                formatted = [
                    datetime.fromisoformat(ts.replace('Z', '+00:00')).strftime('%Y-%m-%d %H:%M:%S')
                    for ts in last_timestamps
                ]
                lines.append("Last 3 message times:")
                lines.extend(f"- {ts}" for ts in formatted)
        else:
            lines.append("⚠️ Channel info: no messages recorded in this channel.")

        await ctx.send("\n".join(lines))

        # Also show general test results
        await ctx.send(f"✅ Test completed! The command `!filter message>#{min_messages} #{channel_name} 24h` {'will work' if filter_24h_pass else 'will not work'}")

    except Exception as e:
        await ctx.send(f"❌ Error during test: {e}")
        logger.error(f"Error in test_channel_filter: {e}")

@bot.command()
async def filter(ctx, *, command_text: str = None):
    """Advanced filtering with Excel export
    Usage: !filter message>10 content>2 @rol1 and @rol2 #kanal 7d
    Example: !filter message>10 content>2 @early agi and @advanced agi #general 7d"""
    try:
        if not command_text:
            await ctx.send("""**Filter Command Help:**

**Syntax:** `!filter [message>count] [content>count] [@role1 and/or @role2] [nohave @role3] [#channel] [time]`

**Examples:**
• `!filter message>10 content>2 @early agi and @advanced agi #general 7d`
• `!filter message>5 content>1 @moderator #help 30d`
• `!filter content>3 @admin or @moderator 1w`
• `!filter #general message>20 content>5`
• `!filter message>15 @vip and @premium or @support 30d`
• `!filter @magnitude 6.0 nohave @magnitude 7.0`
• `!filter @admin nohave @banned 7d`
• `!filter @vip or @premium nohave @suspended`

**Filters:**
• `message>X`: Users with more than X messages
• `content>X`: Users with more than X content links
• `@rol1 and @rol2`: Users with BOTH roles
• `@rol1 or @rol2`: Users with EITHER role
• `nohave @rol3`: Users WITHOUT this role
• `#kanal`: Messages in specific channel
• `7d/1w/1m`: Last 7 days/1 week/1 month

**Operators:** `>`, `<`, `>=`, `<=`, `==`, `!=`

**Results:** Excel file with filtered users""")
            return
        
        # Parse command
        filters = filter_engine.parse_filter_command(command_text)
        
        # Load data
        existing_data = await get_cached_data()
        user_count = sum(1 for _ in iter_user_entries(existing_data))

        if user_count == 0:
            await ctx.send("No user data found.")
            return
        
        # Apply filters
        logger.info(f"Starting filter application with data: {user_count} users")
        logger.info(f"Filters: {filters}")

        # Convert channel IDs to names if needed
        if filters.get('channels'):
            converted_channels = []
            for channel in filters['channels']:
                if channel.isdigit():
                    # Channel ID - convert to name
                    try:
                        channel_obj = ctx.guild.get_channel(int(channel))
                        if channel_obj:
                            converted_channels.append(channel_obj.name)
                        else:
                            converted_channels.append(channel)  # fallback to ID
                    except:
                        converted_channels.append(channel)
                else:
                    converted_channels.append(channel)
            filters['channels'] = converted_channels
            logger.info(f"Converted channel filters: {filters['channels']}")

        results = filter_engine.apply_filters(existing_data, filters)
        logger.info(f"Filter completed. Results: {len(results)} users")
        
        if not results:
            await ctx.send(f"No users found matching the filter: `{command_text}`")
            return
        
        # Generate Excel file directly
        logger.info(f"Generating Excel file for {len(results)} users...")
        excel_file = filter_engine.export_to_excel(results, filters)

        # Log file info
        file_size = os.path.getsize(excel_file) if os.path.exists(excel_file) else 0
        logger.info(f"Generated Excel file: {excel_file}, Size: {file_size} bytes, Results: {len(results)}")

        # Send Excel file
        try:
            logger.info(f"Attempting to send Excel file: {excel_file}")
            logger.info(f"File exists: {os.path.exists(excel_file)}")
            logger.info(f"File size: {os.path.getsize(excel_file) if os.path.exists(excel_file) else 0} bytes")
            logger.info(f"Channel ID: {ctx.channel.id}")
            logger.info(f"Channel name: {ctx.channel.name}")
            logger.info(f"Guild ID: {ctx.guild.id}")

            # Check bot permissions
            bot_permissions = ctx.channel.permissions_for(ctx.guild.me)
            logger.info(f"Bot permissions: {bot_permissions}")
            logger.info(f"Can send messages: {bot_permissions.send_messages}")
            logger.info(f"Can attach files: {bot_permissions.attach_files}")

            # Create and send file
            discord_file = discord.File(
                excel_file,
                filename=f"filter_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            )
            logger.info(f"Discord File object created successfully")

            await ctx.send(file=discord_file)
            logger.info("Excel file sent successfully!")

        except discord.Forbidden as e:
            logger.error(f"Discord Forbidden error: {e}")
            logger.error(f"Bot missing permissions in channel {ctx.channel.name}")
        except discord.HTTPException as e:
            logger.error(f"Discord HTTP error: {e}")
            logger.error(f"Status code: {e.status}")
            logger.error(f"Response: {e.response}")
        except Exception as e:
            logger.error(f"Unexpected error sending file: {e}")
            logger.error(f"Error type: {type(e).__name__}")
            raise
        finally:
            # Clean up temp file
            try:
                os.unlink(excel_file)
            except:
                pass

    except Exception as e:
        logger.error(f"Error in filter command: {e}")
        logger.error(f"Error type: {type(e).__name__}")
        logger.error(f"Error details: {str(e)}")
        await ctx.send(f"Error processing filter: {e}")

@bot.command()
async def debug_content_channels(ctx):
    """Debug content channels - show current configuration"""
    try:
        global content_channels

        await ensure_cache_loaded()

        # Force reload content channels
        load_content_channels()

        # Get current channel info
        current_channel = ctx.channel
        current_channel_id = str(current_channel.id)
        is_content = is_content_channel(current_channel_id)

        storage_info = "messages.json (__content_channels__)"
        last_updated = 'Unknown'
        async with get_message_cache_lock():
            storage_entry = message_cache.get(CONTENT_CHANNELS_STORAGE_KEY, {})
            if isinstance(storage_entry, dict):
                last_updated = storage_entry.get('last_updated', 'Unknown')

        lines = [
            "🔍 Content Channels Debug",
            f"Current channel: {current_channel.name} ({current_channel_id})",
            f"Is content channel: {'Yes' if is_content else 'No'}",
            f"Tracked channel count: {len(content_channels)}",
            f"Storage: {storage_info}",
            f"Last updated: {last_updated}"
        ]

        if content_channels:
            lines.append("Tracked channels:")
            lines.extend(f"- {channel_id}" for channel_id in sorted(content_channels))
        else:
            lines.append("Tracked channels: none")

        if LEGACY_CONTENT_CHANNELS_FILE.exists():
            lines.append("Legacy file detected: content_channels.json (no longer updated)")

        # Split output if necessary
        messages = []
        current = []
        current_len = 0
        for line in lines:
            line_len = len(line) + 1
            if current and current_len + line_len > 1900:
                messages.append("\n".join(current))
                current = ["(cont.)"]
                current_len = len(current[0])
            current.append(line)
            current_len += line_len

        if current:
            messages.append("\n".join(current))

        for message in messages:
            await ctx.send(message)

    except Exception as e:
        await ctx.send(f"Debug error: {e}")
        logger.error(f"Debug content channels error: {e}")


@bot.command()
async def export_raw(ctx):
    """Export all user data to Excel"""
    try:
        await ctx.send("Exporting all user data to Excel...")
        
        # Load data
        existing_data = await get_cached_data()
        entries = list(iter_user_entries(existing_data))

        if not entries:
            await ctx.send("No user data found.")
            return
        
        # Convert to results format
        all_results = []
        for user_id, user_data in entries:
            all_results.append({
                'user_id': user_id,
                'username': user_data.get('username', 'Unknown'),
                'roles': user_data.get('roles', []),
                'total_messages': len(user_data.get('messages', [])),
                'messages': user_data.get('messages', []),
                'first_message': user_data.get('messages', [{}])[0] if user_data.get('messages') else None,
                'last_message': user_data.get('messages', [-1])[-1] if user_data.get('messages') else None
            })
        
        # Sort by message count
        all_results.sort(key=lambda x: x['total_messages'], reverse=True)
        
        # Generate Excel file
        excel_file = filter_engine.export_to_excel(all_results, {})
        
        # Send Excel file
        await ctx.send(
            file=discord.File(
                excel_file,
                filename=f"all_users_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            )
        )
        
        # Clean up temp file
        try:
            os.unlink(excel_file)
        except:
            pass
        
        await ctx.send(f"✅ Exported {len(all_results)} users to Excel")
        
    except Exception as e:
        logger.error(f"Error in export_raw command: {e}")
        await ctx.send(f"Error exporting data: {e}")

@bot.event
async def on_error(event, *args, **kwargs):
    """Global error handler"""
    logger.error(f"Error in event {event}: {args}")

@bot.event
async def on_command_error(ctx, error):
    """Command error handler"""
    if isinstance(error, commands.CommandNotFound):
        return
    if isinstance(error, commands.CheckFailure):
        return
    if isinstance(error, commands.MissingRequiredArgument):
        await ctx.send(f'❌ Missing required argument: `{error.param.name}`')
        return
    logger.error(f"Command error: {error}")

if __name__ == "__main__":
    # Get token from environment
    TOKEN = os.getenv('DISCORD_TOKEN')
    if not TOKEN:
        logger.error("DISCORD_TOKEN not found in .env file!")
        exit(1)

    # Run bot
    bot.run(TOKEN)
